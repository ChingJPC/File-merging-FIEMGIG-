import os
import pandas as pd
from datetime import datetime
import xmltodict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from customtkinter import CTk, CTkFrame, CTkEntry, CTkLabel, CTkButton, CTkCheckBox, CTkFont, CTkImage, CTkProgressBar, StringVar
import customtkinter as ctk
from tkinter import PhotoImage, filedialog, messagebox, Frame, Label, Tk, Text, ttk, filedialog
from PIL import Image, ImageTk
import tkinter as tk
import threading
import sys
import time

def load_custom_font():
    try:
        ctk.load_font("Bangers", "fonts/Bangers.ttf")  # Ruta al archivo de la fuente
    except Exception as e:
        print(f"No se pudo cargar la fuente Bangers: {e}")

# Función para leer archivos Excel con diferentes encabezados y agregar una columna con el nombre del archivo
def read_excel_with_header_and_filename(file_path, header, filename):
    try:
        df = pd.read_excel(file_path, header=header)
        df.insert(0, 'Identificador', os.path.splitext(filename)[0])
        return df
    except Exception as e:
        print(f"Error al leer el archivo {file_path}: {e}")
        return pd.DataFrame()

# Configuracion para mostrar todas las columnas
pd.set_option('display.max_columns', None)

def crear_carpetas():
    # Rutas de las carpetas que se van a crear
    rutas = [
        "C:/file_merging/P04/P04_TODOS_DATOS",
        "C:/file_merging/P04/P04_CURSOS_EVENTOS",
        "C:/file_merging/P04/P04_INTEGRACION",
        "C:/file_merging/P04/P04_FINAL",
        "C:/file_merging/Aprendices/Apre_C_R",
        "C:/file_merging/Aprendices/Apre_Completos",
        "C:/file_merging/Juicios/Juic_C_R",
        "C:/file_merging/Juicios/Juic_Completos",
        "C:/file_merging/Informe"
    ]

    # Crear las carpetas si no existen
    for ruta in rutas:
        os.makedirs(ruta, exist_ok=True)

# Función para convertir archivos XML a DataFrame
def xml_to_df(xml):
    with open(xml, 'r', encoding='utf-8') as f:
        xml_data = f.read().replace("&", "&amp;")
    xml_dict = xmltodict.parse(xml_data)

    print("Convirtiendo archivos XML a XLS...")
    time.sleep(2)  # Simulación de tiempo de conversión
    
    rows = []
    for x in xml_dict["Workbook"]["Worksheet"]["Table"]["Row"][4:]:
        new_row = []
        for old_row in x["Cell"]:
            if not old_row["Data"]:
                continue
            if "#text" not in old_row["Data"]:
                new_row.append(None)
                continue
            new_row.append(old_row["Data"]["#text"])
        rows.append(new_row)
    
    df = pd.DataFrame(rows, columns=rows[0]).iloc[1:]
    return df

# Función para convertir archivos XML a Excel usando xml_to_df
def convert_xml_to_xls(ruta_carpeta, destino_carpeta):
    files = [f for f in os.listdir(ruta_carpeta) if f.endswith('.xml')]

    for f in files:
        file_path = os.path.join(ruta_carpeta, f)
        try:
            df = xml_to_df(file_path)
            nombre_archivo = os.path.splitext(f)[0] + '.xlsx'
            ruta_guardado = os.path.join(destino_carpeta, nombre_archivo)
            df.to_excel(ruta_guardado, index=False)
            print(f"Archivo XML convertido y guardado en: {ruta_guardado}")
        except Exception as e:
            print(f"Error al convertir el archivo {file_path}: {e}")

proceso_pe04_completado = False
proceso_juicio_completado = False

def actualizar_estados_botones():
    global proceso_pe04_completado
    if proceso_pe04_completado:
        button2.configure(state='normal')
        button3.configure(state='normal')
    else:
        button2.configure(state='disabled')
        button3.configure(state='disabled')

# Función para ordenar archivos por fecha, eliminar duplicados y filtrar registros
def procesar_archivos_p04(carpeta_origen):
    global proceso_pe04_completado
    carpeta_destino = "C:/file_merging/P04/P04_TODOS_DATOS"

    # Convertir archivos XML a Excel antes de procesar
    convert_xml_to_xls(carpeta_origen, carpeta_destino)

    archivos = [os.path.join(carpeta_destino, f) for f in os.listdir(carpeta_destino) if f.endswith('.xlsx')]

    archivos_validos = []
    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo)
        if nombre_archivo == 'archivofinal_PE04.xlsx':
            continue

        try:
            fecha_str = nombre_archivo.split('_')[-1].split('.')[0]
            
            if fecha_str.isdigit():
                if len(fecha_str) == 4:
                    fecha = datetime.strptime(fecha_str, '%Y')
                else:
                    fecha = int(fecha_str)
            else:
                raise ValueError("Formato de fecha no reconocido")
            
            archivos_validos.append((fecha, archivo))
        except ValueError:
            print(f"Archivo con nombre inválido para el formato de fecha: {archivo}")

    archivos_ordenados = sorted(archivos_validos, key=lambda x: x[0], reverse=True)

    dfs = []
    total_archivos = len(archivos_ordenados)
    for i, (_, archivo) in enumerate(archivos_ordenados):
        try:
            df = pd.read_excel(archivo)
            dfs.append(df)
        except Exception as e:
            print(f"Error al procesar el archivo {archivo}: {e}")

        # Actualizar barra de progreso
        progress_bar['value'] = (i + 1) / total_archivos * 100
        root.update_idletasks()
        time.sleep(0.1)  # Simulación de tiempo de procesamiento

    if dfs:
        combined_frame = pd.concat(dfs, ignore_index=True)
        combined_frame = combined_frame.drop_duplicates(subset=["IDENTIFICADOR_FICHA"])

        combined_frame = combined_frame[~combined_frame["AMPLIACION_COBERTURA"].str.contains("S", na=False)]

        ruta_destino = os.path.join(carpeta_destino, 'archivofinal_PE04.xlsx')
        combined_frame.to_excel(ruta_destino, index=False)
        print(f"Archivos combinados y guardados en: {ruta_destino}")

        cursos_especiales_eventos = combined_frame[combined_frame["NIVEL_FORMACION"].str.contains("CURSO ESPECIAL|EVENTO", na=False, case=False)]
        cursos_especiales_eventos.to_excel("C:/file_merging/P04/P04_CURSOS_EVENTOS/P04_cursos_eventos.xlsx", index=False)

        integracion = combined_frame[combined_frame["NOMBRE_PROGRAMA_ESPECIAL"].str.startswith("INTEGRACIÓN", na=False)]
        integracion.to_excel("C:/file_merging/P04/P04_INTEGRACION/P04_Integracion.xlsx", index=False)

        restantes = combined_frame[
            ~combined_frame["NIVEL_FORMACION"].str.contains("CURSO ESPECIAL|EVENTO", na=False, case=False) &
            ~combined_frame["NOMBRE_PROGRAMA_ESPECIAL"].str.startswith("INTEGRACIÓN", na=False)
        ]
        restantes.to_excel("C:/file_merging/P04/P04_FINAL/P04_final.xlsx", index=False)

        proceso_pe04_completado = True
        messagebox.showinfo("Proceso PE04", "Proceso PE04 completado correctamente.")

        actualizar_estados_botones()

    else:
        print("No hay objetos para concatenar. La lista 'dfs' está vacía.")

def iniciar_proceso(carpeta_origen):
    progress_bar['value'] = 0
    threading.Thread(target=procesar_archivos_p04, args=(carpeta_origen,)).start()


# Función para leer archivos Excel con diferentes encabezados y agregar una columna con el nombre del archivo
def read_excel_with_header_and_filename(file_path, filename):
    try:
        # Leer solo la fila A5 como encabezado
        df = pd.read_excel(file_path, header=4)  # Asumiendo que A5 es la quinta fila
        df.insert(0, 'Identificador', os.path.splitext(filename)[0])
        return df
    except Exception as e:
        print(f"Error al leer el archivo {file_path}: {e}")
        return pd.DataFrame()

# Función para procesar y combinar archivos de aprendices
def procesar_aprendices(carpeta_aprendices):
    global proceso_pe04_completado
    
    # Verificar si el proceso PE04 ha sido completado
    if not proceso_pe04_completado:
        messagebox.showwarning("Proceso PE04", "Debe completar primero el proceso PE04.")
        return

    if not carpeta_aprendices:
        print("Por favor selecciona una carpeta de aprendices.")
        return

    # Obtener la lista de archivos Excel (.xlsx y .xls) en la carpeta de aprendices
    archivos_aprendices = [
        os.path.join(carpeta_aprendices, archivo) for archivo in os.listdir(carpeta_aprendices)
        if archivo.endswith('.xlsx') or archivo.endswith('.xls')
    ]

    if archivos_aprendices:
        dfs = []
        total_archivos = len(archivos_aprendices)
        for i, archivo in enumerate(archivos_aprendices):
            try:
                if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
                    # Llama a la función para leer el archivo Excel con encabezados en la fila 4 (A5 en Excel)
                    df = read_excel_with_header_and_filename(archivo, filename=os.path.basename(archivo))
                    dfs.append(df)
                else:
                    print(f"Archivo no soportado: {archivo}. Solo se admiten archivos .xlsx y .xls.")
            except Exception as e:
                print(f"Error al procesar el archivo {archivo}: {e}")

            # Actualizar barra de progreso
            progress_bar_aprendices['value'] = (i + 1) / total_archivos * 100
            root.update_idletasks()

        if dfs:
            # Combinar todos los DataFrames en uno solo
            combined_frame = pd.concat(dfs, ignore_index=True)

            # Guardar el archivo combinado en la carpeta de aprendices/Apre_Completos
            ruta_salida_completos = "C:/file_merging/Aprendices/Apre_Completos"
            if not os.path.exists(ruta_salida_completos):
                os.makedirs(ruta_salida_completos)
            
            archivo_combinado = os.path.join(ruta_salida_completos, "DatosCombinadosAprendices.xlsx")
            combined_frame.to_excel(archivo_combinado, index=False)
            print(f"Archivos de aprendices combinados guardados en: {archivo_combinado}")

            # Filtrar y guardar por estado CANCELADO o RETIRO VOLUNTARIO
            df_cancelados = combined_frame[combined_frame['Estado'].isin(['CANCELADO', 'RETIRO VOLUNTARIO','TRASLADADO' ])]
            ruta_salida_cancelados = "C:/file_merging/Aprendices/Apre_C_R"
            if not os.path.exists(ruta_salida_cancelados):
                os.makedirs(ruta_salida_cancelados)
            
            archivo_cancelados = os.path.join(ruta_salida_cancelados, "Apre_Retirados.xlsx")
            df_cancelados.to_excel(archivo_cancelados, index=False)
            print(f"Archivos de aprendices cancelados guardados en: {archivo_cancelados}")

            # Filtrar y guardar por estado CERTIFICADO o TRASLADADO
            df_certificados = combined_frame[combined_frame['Estado'].isin(['CERTIFICADO'])]
            ruta_salida_certificados = "C:/file_merging/Aprendices/Apre_C_R"
            if not os.path.exists(ruta_salida_certificados):
                os.makedirs(ruta_salida_certificados)
            
            archivo_certificados = os.path.join(ruta_salida_certificados, "Apre_Certificados.xlsx")
            df_certificados.to_excel(archivo_certificados, index=False)
            print(f"Archivos de aprendices certificados guardados en: {archivo_certificados}")
            
            messagebox.showinfo("Proceso Aprendices", "Proceso Aprendices completado correctamente.")

        else:
            print("No se encontraron archivos válidos para combinar en la carpeta de aprendices.")
    else:
        print("No se encontraron archivos .xlsx o .xls en la carpeta de aprendices.")

def iniciar_proceso_aprendices(carpeta_aprendices):
    progress_bar_aprendices['value'] = 0
    threading.Thread(target=procesar_aprendices, args=(carpeta_aprendices,)).start()
    actualizar_barra_progreso_aprendices()

def actualizar_barra_progreso_aprendices():
    if threading.active_count() > 1:  # Si el proceso aún se está ejecutando
        progress_bar_aprendices.update_idletasks()
        root.after(100, actualizar_barra_progreso_aprendices)      

        # Función para leer archivos Excel con diferentes encabezados y agregar una columna con el nombre del archivo
def read_excel_with_header_and_filename_juicios(file_path, filename):
    try:
        # Leer solo la fila A5 como encabezado
        df = pd.read_excel(file_path, header=12)  # Asumiendo que A5 es la quinta fila
        df.insert(0, 'Identificador', os.path.splitext(filename)[0])
        return df
    except Exception as e:
        print(f"Error al leer el archivo {file_path}: {e}")
        return pd.DataFrame()
    

def actualizar_estado_boton_informe():
    global proceso_juicio_completado
    if proceso_juicio_completado:
        button4.configure(state='normal')
    else:
        button4.configure(state='disabled')  

# Función para procesar y combinar archivos de juicios
def procesar_juicios(carpeta_juicios):
    global proceso_juicio_completado
    
    # Verificar si el proceso PE04 ha sido completado
    if not proceso_pe04_completado:
        messagebox.showwarning("Proceso PE04", "Debe completar primero el proceso PE04.")
        return
    
    if not carpeta_juicios:
        print("Por favor selecciona una carpeta de juicios.")
        return

    # Obtener la lista de archivos Excel (.xlsx y .xls) en la carpeta de juicios
    archivos_juicios = [
        os.path.join(carpeta_juicios, archivo) for archivo in os.listdir(carpeta_juicios)
        if archivo.endswith('.xlsx') or archivo.endswith('.xls')
    ]

    if archivos_juicios:
        dfs = []
        total_archivos = len(archivos_juicios)
        for i, archivo in enumerate(archivos_juicios):
            try:
                if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
                    # Llama a la función para leer el archivo Excel con encabezados en la fila 11 (A12 en Excel)
                    df = read_excel_with_header_and_filename_juicios(archivo, filename=os.path.basename(archivo))
                    dfs.append(df)
                else:
                    print(f"Archivo no soportado: {archivo}. Solo se admiten archivos .xlsx y .xls.")
            except Exception as e:
                print(f"Error al procesar el archivo {archivo}: {e}")

            # Actualizar barra de progreso
            progress_bar_juicios['value'] = (i + 1) / total_archivos * 100
            root.update_idletasks()

        if dfs:
            # Combinar todos los DataFrames en uno solo
            combined_frame = pd.concat(dfs, ignore_index=True)

            # Guardar el archivo combinado en la carpeta de juicios/Juic_Completos
            ruta_salida_completos = "C:/file_merging/Juicios/Juic_Completos"
            if not os.path.exists(ruta_salida_completos):
                os.makedirs(ruta_salida_completos)
            
            archivo_combinado = os.path.join(ruta_salida_completos, "DatosCombinadosJuicios.xlsx")
            combined_frame.to_excel(archivo_combinado, index=False)
            print(f"Archivos de juicios combinados guardados en: {archivo_combinado}")

            # Filtrar y guardar por estado CANCELADO, RETIRO VOLUNTARIO Y TRASLADADO
            df_cancelados = combined_frame[combined_frame['Estado'].isin(['CANCELADO', 'RETIRO VOLUNTARIO', 'TRASLADADO'])]
            ruta_salida_cancelados = "C:/file_merging/Juicios/Juic_C_R"
            if not os.path.exists(ruta_salida_cancelados):
                os.makedirs(ruta_salida_cancelados)
            
            archivo_cancelados = os.path.join(ruta_salida_cancelados, "Juic_Retirados.xlsx")
            df_cancelados.to_excel(archivo_cancelados, index=False)
            print(f"Archivos de juicios cancelados guardados en: {archivo_cancelados}")

            # Filtrar y guardar por estado CERTIFICADO o TRASLADADO
            df_certificados = combined_frame[combined_frame['Estado'].isin(['CERTIFICADO'])]
            ruta_salida_certificados = "C:/file_merging/Juicios/Juic_C_R"
            if not os.path.exists(ruta_salida_certificados):
                os.makedirs(ruta_salida_certificados)
            
            archivo_certificados = os.path.join(ruta_salida_certificados, "Juic_Certificados.xlsx")
            df_certificados.to_excel(archivo_certificados, index=False)
            print(f"Archivos de juicios certificados guardados en: {archivo_certificados}")

            proceso_juicio_completado = True
            messagebox.showinfo("Proceso Juicios", "Proceso Juicios completado correctamente.")

            actualizar_estado_boton_informe()

        else:
            print("No se encontraron archivos válidos para combinar en la carpeta de juicios.")
    else:
        print("No se encontraron archivos .xlsx o .xls en la carpeta de juicios.")

def iniciar_proceso_juicios(carpeta_juicios):
    progress_bar_juicios['value'] = 0
    threading.Thread(target=procesar_juicios, args=(carpeta_juicios,)).start()
    actualizar_barra_progreso_juicios()

def actualizar_barra_progreso_juicios():
    if threading.active_count() > 1:  # Si el proceso aún se está ejecutando
        progress_bar_juicios.update_idletasks()
        root.after(600, actualizar_barra_progreso_juicios)

#Crear Informe
def generar_informe():
    ruta_guardado = "C:/file_merging/Informe/informe_generado.xlsx"
    os.makedirs(os.path.dirname(ruta_guardado), exist_ok=True)

    def proceso():
        try:
            # Cargar datos de los archivos
            juic_certificados = pd.read_excel("C:/file_merging/Juicios/Juic_C_R/Juic_Certificados.xlsx")
            p04_completo = pd.read_excel("C:/file_merging/P04/P04_FINAL/P04_final.xlsx")
            
            # Crear un nuevo archivo Excel con openpyxl
            workbook = Workbook()
            
            # Crear la hoja "Informe" y agregar los encabezados en negrita
            hoja_informe = workbook.active
            hoja_informe.title = "Informe"

            encabezados_informe = [
                "Ficha&ID", "Técnica", "Transversal", "Ficha", 
                "Lider", "Programa", "Titulo", "Identificación", 
                "Nombre", "Apellidos", "Tipo de Documento"
            ]
            
            for col, encabezado in enumerate(encabezados_informe, start=1):
                cell = hoja_informe.cell(row=1, column=col, value=encabezado)
                cell.font = Font(bold=True)

            # Crear la hoja "Juic_Certificados" y copiar los datos
            hoja_juic_certificados = workbook.create_sheet("Juic_Certificados")
            hoja_juic_certificados.append(juic_certificados.columns.tolist())  # Encabezados
            for col in hoja_juic_certificados.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(juic_certificados.columns)):
                for cell in col:
                    cell.font = Font(bold=True)
            for _, row in juic_certificados.iterrows():
                hoja_juic_certificados.append(row.tolist())

            # Crear la hoja "P04_Completo" y copiar los datos
            hoja_p04_completo = workbook.create_sheet("P04_Completo")
            hoja_p04_completo.append(p04_completo.columns.tolist())  # Encabezados
            for col in hoja_p04_completo.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(p04_completo.columns)):
                for cell in col:
                    cell.font = Font(bold=True)
            for _, row in p04_completo.iterrows():
                hoja_p04_completo.append(row.tolist())

            # Crear la hoja "Ficha&Ambiente" y agregar los encabezados en negrita
            hoja_ficha_ambiente = workbook.create_sheet("Ficha&Ambiente")

            encabezados_ficha_ambiente = [
                "FICHA APRENDICES", "FICHA", "CODIGO", "VR*SENA", "OFERTA", "TITULO", 
                "NOMBRE DEL PROGRAMA", "LIDER", "MATRICULAS", "ACTIVOS", "MESES LECTIVA", 
                "INICIO LECTIVA", "FINAL LECTIVA", "PORCENTAJE DE AVANCE", 
                "AVANCE JUICIOS TECNICA", "FALTA POR EVALUAR", 
                "AVANCE JUICIOS TRANSVERSALES", "FALTA POR EVALUAR", 
                "EVALUACION TOTAL", "MESES PRODUCTIVA", "INICIO PRODUCTIVA", 
                "FINAL PRODUCTIVA", "COORDINADOR", "ESTADO CURSO", 
                "MODALIDAD FORMACION", "CODIGO NIVEL DE FORMACION", "RESPUESTA", 
                "PROGRAMA ESPECIAL", "DESERCION", "ACTIVOS REAL", "PROYECTO"
            ]
            hoja_ficha_ambiente.append(encabezados_ficha_ambiente)

            # Establecer el color verde claro en los encabezados relevantes
            verde_claro = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            columnas_verde = [2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 24, 25, 26, 28]
            for col in columnas_verde:
                hoja_ficha_ambiente.cell(row=1, column=col).fill = verde_claro
                hoja_ficha_ambiente.cell(row=1, column=col).font = Font(bold=True)

            # Establecer el color rojo y negrita para los encabezados
            rojo_fuerte = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            columnas_rojas = [11, 13, 14, 21, 22, 23]  # Columnas a aplicar el formato (Final Lectiva, Porcentaje de Avance, Meses Productiva, Inicio Productiva, Final Productiva)
            for col in columnas_rojas:
                hoja_ficha_ambiente.cell(row=1, column=col).fill = rojo_fuerte
                hoja_ficha_ambiente.cell(row=1, column=col).font = Font(bold=True)

            # Transferir datos desde P04_Completo a Ficha&Ambiente
            for index, row in p04_completo.iterrows():
                hoja_ficha_ambiente.append([
                    "",  # FICHA APRENDICES
                    row['IDENTIFICADOR_FICHA'],  # FICHA
                    row['CODIGO_PROGRAMA'],  # CODIGO
                    row['VERSION_PROGRAMA'],  # VR*SENA
                    "",  # OFERTA (vacío)
                    row['NIVEL_FORMACION'],  # Titulo
                    row['NOMBRE_PROGRAMA_FORMACION'],  # NOMBRE DEL PROGRAMA
                    row['NOMBRE_RESPONSABLE'],  # LIDER
                    row['TOTAL_APRENDICES'],  # MATRICULAS
                    row['TOTAL_APRENDICES_ACTIVOS'],  # ACTIVOS
                    "",  # MESES LECTIVA (Vacio)
                    row['FECHA_INICIO_FICHA'],  # INICIO LECTIVA
                    "",  # FINAL LECTIVA (vacío)
                    "",  # PORCENTAJE DE AVANCE (vacío)
                    "",  # AVANCE JUICIOS TECNICA (vacío)
                    "",  # FALTA POR EVALUAR (vacío)
                    "",  # AVANCE JUICIOS TRANSVERSALES (vacío)
                    "",  # FALTA POR EVALUAR (vacío)
                    "",  # EVALUACION TOTAL (vacío)
                    "",  # MESES PRODUCTIVA (vacío)
                    "",  # INICIO PRODUCTIVA (vacío)
                    "",  # FINAL PRODUCTIVA (vacío)
                    "",  # COORDINADOR (vacío)
                    row['ESTADO_CURSO'],  # ESTADO CURSO
                    row['MODALIDAD_FORMACION'],  # MODALIDAD FORMACION
                    row['CODIGO_NIVEL_FORMACION'],  # CODIGO NIVEL DE FORMACION
                    "",  # RESPUESTA (vacío)
                    row['NOMBRE_PROGRAMA_ESPECIAL'],  # PROGRAMA ESPECIAL
                    "",  # DESERCION (vacío)
                    "",  # ACTIVOS REAL (vacío)
                    ""  # PROYECTO (vacío)
                ])

            for row_num in range(2, hoja_ficha_ambiente.max_row + 1):

                # Añadir la fórmula en la columna "FINAL LECTIVA"
                u_cell = hoja_ficha_ambiente.cell(row=row_num, column=21).coordinate
                hoja_ficha_ambiente.cell(row=row_num, column=13, value=f"={u_cell}-1")

                # Añadir la fórmula en la columna "MESES LECTIVA"
                u_cell = hoja_ficha_ambiente.cell(row=row_num, column=12).coordinate
                l_cell = hoja_ficha_ambiente.cell(row=row_num, column=13).coordinate
                hoja_ficha_ambiente.cell(row=row_num, column=11, value=f"=ROUND(DAYS({u_cell},{l_cell})/30,0)")

                # Añadir la fórmula en la columna "MESES PRODUCTIVA"
                z_cell = hoja_ficha_ambiente.cell(row=row_num, column=26).coordinate
                hoja_ficha_ambiente.cell(row=row_num, column=21, value=f"=IF({z_cell}=6,6,IF({z_cell}=2,6,IF({z_cell}=10,3,0)))")

                # Añadir la fórmula en la columna "INICIO PRODUCTIVA"
                v_cell = hoja_ficha_ambiente.cell(row=row_num, column=22).coordinate
                t_cell = hoja_ficha_ambiente.cell(row=row_num, column=21).coordinate
                hoja_ficha_ambiente.cell(row=row_num, column=22, value=f"=EDATE({v_cell},-{t_cell})")

                # Añadir la fórmula en la columna "PORCENTAJE DE AVANCE"
                l_cell = hoja_ficha_ambiente.cell(row=row_num, column=12).coordinate
                m_cell = hoja_ficha_ambiente.cell(row=row_num, column=13).coordinate
                hoja_ficha_ambiente.cell(row=row_num, column=14, value=f"=IF(DAYS360({l_cell},TODAY())/DAYS360({l_cell},{m_cell})>1,1,DAYS360({l_cell},TODAY())/DAYS360({l_cell},{m_cell}))")

                # Añadir la fórmula en la columna "FINAL PRODUCTIVA"
                hoja_ficha_ambiente.cell(row=row_num, column=23, value=f"=VLOOKUP($E{row_num},'P04_Completo'!$E:$AZ,V$1,FALSE)")

            progress_bar['value'] = 100
            progress_bar.update_idletasks()

            # Guardar el archivo
            workbook.save(ruta_guardado)

            # Mostrar mensaje de éxito
            messagebox.showinfo("Generación de Informe", "Generación de Informe completado correctamente.")
        
        except Exception as e:
            messagebox.showerror("Error", f"Hubo un error al generar el informe: {e}")

    # Iniciar el proceso en un hilo separado
    threading.Thread(target=proceso).start()


# Creación de carpetas al inicio del programa
crear_carpetas()
        
# Función para seleccionar carpeta 
def seleccionar_carpeta(entry):
    carpeta = filedialog.askdirectory()
    entry.delete(0, 'end')
    entry.insert(0, carpeta)

# Función para seleccionar la carpeta de aprendices
def seleccionar_carpeta_aprendices(entry_aprendices):
    carpeta_seleccionada = filedialog.askdirectory()
    if carpeta_seleccionada:
        entry_aprendices.set(carpeta_seleccionada)

# Función para seleccionar la carpeta de juicios
def seleccionar_carpeta_juicios(entry_juicios):
    carpeta_seleccionada = filedialog.askdirectory()
    if carpeta_seleccionada:
        entry_juicios.set(carpeta_seleccionada)

        
# Inicio de interfaz grafica
class CustomApp(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        
        self.vista_actual = None

        # Frame principal para organizar las vistas
        self.frame_principal = ctk.CTkFrame(self, fg_color='white')
        self.frame_principal.pack(fill=tk.BOTH, expand=True)

        # Vistas
        self.vista1 = self.crear_vista("ARCHIVOS P04", "- En los archivos P04:\n\n"
                                         "Procesamiento del P04:\n"
                                         "- Convertir archivos XML a XLS.\n"
                                         "- Ordenar los archivos del P04 de más reciente a más antiguo.\n"
                                         "- Eliminar duplicados basados en el identificador de ficha.\n"
                                         "- Quitar registros con fichas de ampliación de cobertura que contienen la letra 'S'.\n\n"
                                         "Creación de archivos para P04:\n"
                                         "- Un archivo con solo cursos especiales y eventos.\n"
                                         "- Un archivo con solo datos de integración.\n"
                                         "- Un archivo final sin ninguno de los dos tipos anteriores, con duplicados eliminados.")
        
        self.vista2 = self.crear_vista("ARCHIVOS APRENDICES",">>Clasificación de aprendices:\n\n"
                                        "- Combinar un archivo completo.\n"
                                        "- Separar los aprendices en activos y cancelados.")
                                       
        self.vista3 = self.crear_vista("ARCHIVOS JUICIOS", ">>Clasificación de juicios:\n\n"
                                        "- Combinar un archivo completo.\n"
                                        "- Separar los juicios en activos y cancelados.")
        
        self.vista4 = self.crear_vista("ARCHIVO DE INFORME", ">>Informacion sobre el informe:\n\n"
                                        "- Hace un analizis global de todos los archivos.\n"
                                        "- Hace calculos para sacar el porcentje de cierta informacion.")

        # Mostramos la primera vista por defecto
        self.mostrar_vista(self.vista1)
        
    def crear_vista(self, titulo, mensaje):
        frame = ctk.CTkFrame(self.frame_principal, fg_color='white')

        # Crear y configurar la etiqueta con el título
        label = ctk.CTkLabel(frame, text=titulo, font=("Arial", 24), text_color='black')
        label.pack(pady=20)

        # Crear el área de texto para mostrar el mensaje
        text_frame = ctk.CTkFrame(frame, fg_color='white')
        text_frame.pack(pady=0, padx=20)

        label_text = ctk.CTkLabel(text_frame, text=mensaje, wraplength=900, justify=tk.LEFT, text_color='black')
        label_text.pack(pady=0)

        # Crear la barra de copyright
        copyright_bar = ctk.CTkFrame(frame, bg_color='#1FAD00', fg_color='#1FAD00', height=30)
        copyright_bar.pack(side=tk.BOTTOM, fill=tk.X)

        copyright_label = ctk.CTkLabel(copyright_bar, text="© 2024 File Merging - SENA", fg_color='#1FAD00', bg_color='#1FAD00')
        copyright_label.pack(pady=5, padx=10)

        
        return frame

    def mostrar_vista(self, vista):
        # Ocultar la vista actual si hay alguna
        if self.vista_actual:
            self.vista_actual.pack_forget()
        
        # Mostrar la nueva vista
        vista.pack(fill=tk.BOTH, expand=True)
        self.vista_actual = vista
        
    def cambiar_vista(self, vista):
        self.mostrar_vista(vista)

if __name__ == "__main__":
    root = ctk.CTk()
    root.title("File Merging")
    root.geometry("1100x800")
    root.configure(bg='#EFEFEF')
    root.resizable(False, False)


    # Iconos del Software
    logogeneral = PhotoImage(file='imagenes/filemergin.png')

    # Redimensionar la imagen
    new_size = (150, 150)  # Nuevo tamaño (ancho, alto)
    image = Image.open('imagenes/filemergin.png')
    resized_image = image.resize(new_size)
    logogeneral = ImageTk.PhotoImage(resized_image)

    # Crear el frame para el titulo
    frame_logo = ctk.CTkFrame(root, fg_color='#1FAD00', height=150,)
    frame_logo.grid(row=0, column=0, sticky='nwe', pady=(0,8), padx=(8,8))


    # Asignar la imagen al label y expandirlo para ocupar toda la celda del grid
    label_logo = Label(frame_logo, image=logogeneral, bg='#1FAD00')
    label_logo.grid(row=0, column=0, padx=(20,0), sticky='w')

    # Agregar el texto "FILE MERGING"
    title_font = ("Bangers", 48)
    label_texto = Label(frame_logo, text="FILE MERGING - SENA", font=title_font, fg="white", bg='#1FAD00')
    label_texto.grid(row=0, column=1, padx=(200,0), sticky='we')

    # Definir una fuente personalizada
    fuente_subtitulos = ctk.CTkFont(family="Helvetica", size=15, weight="bold")
    fuente_rutas = ctk.CTkFont(family="Helvetica", size=12, slant="italic")

    # Configurar el grid layout en root
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Crear el frame principal para la barra de botones a la izquierda
    frame_botones = ctk.CTkFrame(root, fg_color='white', width=200)
    frame_botones.grid(row=0, column=0, sticky='nsw', pady=(158,8), padx=(8,8))
    
    # Crear el frame principal para las vistas a la derecha
    frame_vistas_derecha = ctk.CTkFrame(root, fg_color='#EFEFEF', bg_color='#EFEFEF', width=100)
    frame_vistas_derecha.grid(row=0, column=0, sticky='nsew', padx=(178,8), pady=(158,8))

    # Crear los botones dentro de frame_botones
    app = CustomApp(frame_vistas_derecha)  # Instanciar CustomApp con el frame_vistas_derecha

    # Botones en frame_botones para cambiar de vista
    button1 = ctk.CTkButton(frame_botones, text="P04", fg_color='#1FAD00', command=lambda: app.cambiar_vista(app.vista1))
    button1.grid(row=0, column=0, padx=10, pady=10, sticky='ew')
    
    button2 = ctk.CTkButton(frame_botones, text="Aprendices", fg_color='#1FAD00', command=lambda: app.cambiar_vista(app.vista2))
    button2.grid(row=1, column=0, padx=10, pady=10, sticky='ew')
        
    button3 = ctk.CTkButton(frame_botones, text="Juicios", fg_color='#1FAD00', command=lambda: app.cambiar_vista(app.vista3))
    button3.grid(row=2, column=0, padx=10, pady=10, sticky='ew')

    button4 = ctk.CTkButton(frame_botones, text="Informe", fg_color='#1FAD00', command=lambda: app.cambiar_vista(app.vista4))
    button4.grid(row=3, column=0, padx=10, pady=10, sticky='ew')

    if not proceso_pe04_completado:
        button2.configure(state='disabled')
        button3.configure(state='disabled')
        button4.configure(state='disabled')

    # Crear widgets  para la primera vista
    frame_widgets_vista1 = ctk.CTkFrame(app.vista1, fg_color='white')
    frame_widgets_vista1.pack(pady=20, padx=20)

    ctk.CTkLabel(frame_widgets_vista1, text="Selecciona la carpeta de PE04: ", text_color="#1FAD00", font=fuente_subtitulos).grid(row=0, column=0, padx=(20,40), pady=(20,20), sticky='E')
    entry_pe04 = ctk.CTkEntry(frame_widgets_vista1, width=350, font=("Arial", 12), border_color='#1B9800', fg_color="#1FAD00")
    entry_pe04.grid(row=0, column=1, padx=(0,40), pady=(20, 20), sticky='w')
    ctk.CTkButton(frame_widgets_vista1, text="Seleccionar", fg_color='#1FAD00', command=lambda: seleccionar_carpeta(entry_pe04)).grid(row=0, column=2, padx=(0,20), pady=(20,20), sticky='W')
    ctk.CTkButton(frame_widgets_vista1, text="Procesar PE04", fg_color='#1FAD00', command=lambda: iniciar_proceso(entry_pe04.get())).grid(row=1, column=2, pady=(20, 20), padx=(0, 20), sticky='e')
    
    # Crear widgets para la segunda vista
    frame_widgets_vista2 = ctk.CTkFrame(app.vista2, fg_color='white')
    frame_widgets_vista2.pack(pady=20, padx=20)

    ctk.CTkLabel(frame_widgets_vista2, text="Selecciona la carpeta de Aprendices: ", text_color="#1FAD00", font=fuente_subtitulos).grid(row=0, column=0, padx=(20,40), pady=(20,20), sticky='E')
    entry_aprendices = ctk.CTkEntry(frame_widgets_vista2, width=350, font=("Arial", 12), border_color='#1B9800', fg_color="#1FAD00")
    entry_aprendices.grid(row=0, column=1, padx=(0,40), pady=(20, 20), sticky='w')
    ctk.CTkButton(frame_widgets_vista2, text="Seleccionar", fg_color='#1FAD00', command=lambda: seleccionar_carpeta(entry_aprendices)).grid(row=0, column=2, padx=(0,20), pady=(20,20), sticky='W')
    ctk.CTkButton(frame_widgets_vista2, text="Procesar Aprendices", fg_color='#1FAD00', command=lambda: iniciar_proceso_aprendices(entry_aprendices.get())).grid(row=1, column=2, pady=(20, 20), padx=(0, 20), sticky='e')

    # Crear widgets para la tercera vista
    frame_widgets_vista3 = ctk.CTkFrame(app.vista3, fg_color='white')
    frame_widgets_vista3.pack(pady=20, padx=20)

    ctk.CTkLabel(frame_widgets_vista3, text="Selecciona la carpeta de Juicios: ", text_color="#1FAD00", font=fuente_subtitulos).grid(row=0, column=0, padx=(20,40), pady=(20,20), sticky='E')
    entry_juicios = ctk.CTkEntry(frame_widgets_vista3, width=350, font=("Arial", 12), border_color='#1B9800', fg_color="#1FAD00")
    entry_juicios.grid(row=0, column=1, padx=(0,40), pady=(20, 20), sticky='w')
    ctk.CTkButton(frame_widgets_vista3, text="Seleccionar", fg_color='#1FAD00', command=lambda: seleccionar_carpeta(entry_juicios)).grid(row=0, column=2, padx=(0,20), pady=(20,20), sticky='W')
    ctk.CTkButton(frame_widgets_vista3, text="Procesar Juicios", fg_color='#1FAD00', command=lambda: iniciar_proceso_juicios(entry_juicios.get())).grid(row=1, column=2, pady=(20, 20), padx=(0, 20), sticky='e')
    
    # Crear widgets para la cuarta vista
    frame_widgets_vista4 = ctk.CTkFrame(app.vista4, fg_color='white')
    frame_widgets_vista4.pack(pady=20, padx=20)
    boton_generar_informe = ctk.CTkButton(frame_widgets_vista4, text="Generar Informe", fg_color='#1FAD00', command=generar_informe)
    boton_generar_informe.grid(row=0, column=0, columnspan=3, pady=(10, 10))

    label_progreso = ctk.CTkLabel(frame_widgets_vista1, text="Barra de progreso", text_color='black')
    label_progreso.grid(row=1, column=0, columnspan=3, pady=(25, 5), sticky='s')
    progress_bar = ttk.Progressbar(frame_widgets_vista1, orient="horizontal", length=600, mode="determinate")
    progress_bar.grid(row=2, column=0, columnspan=3, pady=(5, 20))

    label_progreso_aprendices = ctk.CTkLabel(frame_widgets_vista2, text="Barra de progreso", text_color='black')
    label_progreso_aprendices.grid(row=1, column=0, columnspan=3, pady=(25, 5), sticky='s')
    progress_bar_aprendices = ttk.Progressbar(frame_widgets_vista2, orient="horizontal", length=600, mode="determinate")
    progress_bar_aprendices.grid(row=2, column=0, columnspan=3, pady=(5, 20))

    label_progreso_juicios = ctk.CTkLabel(frame_widgets_vista3, text="Barra de progreso", text_color='black')
    label_progreso_juicios.grid(row=1, column=0, columnspan=3, pady=(25, 5), sticky='s')
    progress_bar_juicios = ttk.Progressbar(frame_widgets_vista3, orient="horizontal", length=600, mode="determinate")
    progress_bar_juicios.grid(row=2, column=0, columnspan=3, pady=(5, 20))

    label_progreso_informe = ctk.CTkLabel(frame_widgets_vista4, text="Barra de progreso", text_color='black')
    label_progreso_informe.grid(row=1, column=0, columnspan=3, pady=(25, 5), sticky='s')
    progress_bar_informe = ttk.Progressbar(frame_widgets_vista4, orient="horizontal", length=600, mode="determinate")
    progress_bar_informe.grid(row=2, column=0, columnspan=3, pady=(5, 20))
    

    # Centrar la ventana en la pantalla
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - 1100) // 2
    y = (screen_height - 800) // 2
    root.geometry("+{}+{}".format(x, y))
    
    load_custom_font()
    app.pack(fill=tk.BOTH, expand=True)
    root.mainloop()